import tkinter as tk
from tkinter import ttk, filedialog,messagebox
from tkinter import *
from tkinter.simpledialog import askstring
from tkinter.messagebox import showinfo
from tkcalendar import *
import os
import cv2
import sys
from PIL import Image, ImageTk
import numpy as np
import pytesseract
import openpyxl
from openpyxl import Workbook
import datetime
import pandas as pd
import serial
import time
import mysql.connector

fileName = os.environ['ALLUSERSPROFILE'] + "\WebcamCap.txt"
cancel = False

#conn= mysql.connector.connect(
    #host='localhost',
    #user='root',
    #database='dbstatusled',
    #password='',
#)

def prompt_ok(event = 0):
    global cancel, button, button1, button2, button3
    cancel = True

    button.place_forget()
    button1 = tk.Button(mainWindow, text="Good Image!", command=save)
    button2 = tk.Button(mainWindow, text="Try Again", command=resume)
    button1.place(anchor=tk.W, relx=0.32, rely=0.1, width=140, height=40)
    button2.place(anchor=tk.W, relx=0.32, rely=0.2, width=140, height=40)
    button1.focus()
    button3 = tk.Button(mainWindow, text="Ocr Sekarang", command=procr)
    button3.place_forget()
    
def save(event = 0):
    global prevImg, button3

    if (len(sys.argv) < 2):
        filepath = "imageCap.png"
    else:
        filepath = sys.argv[1]

    print ("Output file to: " + filepath)
    prevImg.save(filepath)
    
    button3.place(anchor=tk.W, relx=0.32, rely=0.3, width=140, height=40)
    
    
def procr(event = 0):
    global perusahaan,alamat,jenis,nomor,tanggal
    
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    loksam = r"C:\Users\ASUS\Documents\lemari\lemari_project\imageCap.png"
    sampel = cv2.imread(loksam)
    hasiltes = pytesseract.image_to_string(sampel)
    ontk = hasiltes.replace("\n", " ")
    sptc = hasiltes.replace("\n", "#")
    sforc = sptc.replace("##","#")
    var.set(ontk)
    print(ontk)
    perusahaan,alamat,jenis,nomor,nouse = sforc.split('#')
    time = datetime.datetime.now()
    tanggal = time.strftime("%d-%m-%Y")
    print("nama perusahaan :", perusahaan)
    print("alamat perusahaan :", alamat)
    print("jenis arsip :", jenis)
    print("nomor :", nomor)
    print("waktu :", tanggal)
    putin_excel()
    
    
def putin_excel():
    global perusahaan,alamat,jenis,nomor,tanggal,mainWindow,tk,messagebox,conn
    tgl, bln, thn = tanggal.split('-')
    bln2 = ('a'+bln)
    bln3 = ('b'+bln)
    
    if jenis == ("INVOICE") and perusahaan == ("PT. Bintang Utara"):
        keterangan = ("masuk")
        name = ask_nonempty_string("Name", "What is your name?")
        perihal = ask_nonempty_string('Konten Arsip', 'Perihal ?' )
        exfil= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx"
        workbook = openpyxl.load_workbook(exfil)
        sheet = workbook.active
        sheet.append([jenis,nomor,perusahaan,tanggal,keterangan,name,perihal])
        workbook.save(exfil)
        
        exfil2= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book2.xlsx"
        workbook2 = openpyxl.load_workbook(exfil2)
        sheet2 = workbook2.active
        sheet2.append([jenis,nomor,perusahaan,tanggal,keterangan,name,perihal])
        workbook2.save(exfil2)
        
        mycursor= conn.cursor()
        sendbase = "UPDATE statusled SET Stat=%s WHERE ID=0"
        val=(bln,)
        mycursor.execute(sendbase,val)
        conn.commit()
        print("{} Data Berhasil Dimasukan".format(mycursor.rowcount))
        msg_box = tk.messagebox.askquestion('Drawer Check', 'Segera Letakan Arsip Kedalam Map',
                                        icon='warning')
        if msg_box == 'yes':
            sendbase = "UPDATE statusled SET Stat='99' WHERE ID=0"
            mycursor.execute(sendbase)
            conn.commit()
            tk.messagebox.showwarning(title="Success", message="Arsip telah dicatat, Tutup Laci Kembali !!!")
            open_file()
        
    if jenis == ("INVOICE") and perusahaan != ("PT. Bintang Utara") :
        keterangan = ("masuk")
        name = ask_nonempty_string("Name", "What is your name?")
        perihal = ask_nonempty_string('Konten Arsip', 'Perihal ?' )
        exfil= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx"
        workbook = openpyxl.load_workbook(exfil)
        sheet = workbook.active
        sheet.append([jenis,nomor,perusahaan,tanggal,keterangan,name,perihal])
        workbook.save(exfil)
        
        exfil2= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book2.xlsx"
        workbook2 = openpyxl.load_workbook(exfil2)
        sheet2 = workbook2.active
        sheet2.append([jenis,nomor,perusahaan,tanggal,keterangan,name,perihal])
        workbook2.save(exfil2)
        
        mycursor= conn.cursor()
        sendbase = "UPDATE statusled SET Stat=%s WHERE ID=0"
        val=(bln2,)
        mycursor.execute(sendbase,val)
        conn.commit()
        print("{} Data Berhasil Dimasukan".format(mycursor.rowcount))
        msg_box = tk.messagebox.askquestion('Drawer Check', 'Segera Letakan Arsip Kedalam Map !!!',
                                        icon='warning')
        if msg_box == 'yes':
            sendbase = "UPDATE statusled SET Stat='99' WHERE ID=0"
            mycursor.execute(sendbase)
            conn.commit()
            tk.messagebox.showwarning(title="Success", message="Arsip telah dicatat, Tutup Laci Kembali !!!")
            open_file()
        
    if jenis == ("PERIZINAN") :
        keterangan = ("masuk")
        name = ask_nonempty_string("Name", "What is your name?")
        perihal = ask_nonempty_string('Konten Arsip', 'Perihal ?')
        exfil= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx"
        workbook = openpyxl.load_workbook(exfil)
        sheet = workbook.active
        sheet.append([jenis,nomor,perusahaan,tanggal,keterangan,name,perihal])
        workbook.save(exfil)
        
        exfil2= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book2.xlsx"
        workbook2 = openpyxl.load_workbook(exfil2)
        sheet2 = workbook2.active
        sheet2.append([jenis,nomor,perusahaan,tanggal,keterangan,name,perihal])
        workbook2.save(exfil2)
        
        mycursor= conn.cursor()
        sendbase = "UPDATE statusled SET Stat=%s WHERE ID=0"
        val=(bln3,)
        mycursor.execute(sendbase,val)
        conn.commit()
        print("{} Data Berhasil Dimasukan".format(mycursor.rowcount))
        msg_box = tk.messagebox.askquestion('Drawer Check', 'Segera Letakan Arsip Kedalam Map !!!',
                                        icon='warning')
        if msg_box == 'yes':
            sendbase = "UPDATE statusled SET Stat='99' WHERE ID=0"
            mycursor.execute(sendbase)
            conn.commit()
            tk.messagebox.showwarning(title="Success", message="Arsip telah dicatat, Tutup Laci Kembali !!!")
            open_file()
    
    if jenis == ("PAJAK"):
        tk.messagebox.showwarning(title="Error", message="Arsip bukan bagian dari lemari ini")
    
    if jenis == ("PENAWARAN"):
        tk.messagebox.showwarning(title="Error", message="Arsip bukan bagian dari lemari ini")
    

def ask_nonempty_string(title, prompt):
    user_input = askstring(title, prompt)
    while not user_input:
        showinfo("Error", "Input tidak boleh kosong!")
        user_input = askstring(title, prompt)
    return user_input


def open_file():
   exloc = (r'C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx')

   if exloc:
         exloc = r"{}".format(exloc)
         df = pd.read_excel(exloc)

   # Clear all the previous data in tree
   clear_treeview()

   # Add new data in Treeview widget
   tree["column"] = list(df.columns)
   tree["show"] = "headings"

   # For Headings iterate over the columns
   for col in tree["column"]:
        tree.column(col, width=90, anchor="c")
        tree.heading(col, text=col)

   # Put Data in Rows
   df_rows = df.to_numpy().tolist()
   for row in df_rows:
       tree.insert("", "end", values=row)
    
   style = ttk.Style()
   style.theme_use('clam')
   
   tree.place(bordermode=tk.INSIDE, relx=0.0, rely=0.71, anchor=tk.W,width=640,height=310)
   tree_scroll.place(bordermode=tk.INSIDE, relx=0.64, rely=0.71, anchor=tk.W,width=15,height=310,)


def search ():
    global tree
    
    clear_treeview()
    df = pd.read_excel(r'C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx')
    l1 = list(df)  # List of column names as headers
    query = se_entry.get().strip() # get user entered string
    
    str1 = df.Nomor.str.contains(query, case= False) # name column value matching
    str2 = df.Perihal.str.contains(query, case= False)
    str3 = df.Jenis.str.contains(query, case= False)
    str4 = df.Nama.str.contains(query, case= False)
    str5 = df.Perusahaan.str.contains(query, case= False)
    str6 = df.Tanggal.str.contains(query, case= False)
        
    df2 = df[( str1 + str2 + str3 + str4 + str5 + str6 )]  # combine all conditions using | operator
    r_set = df2.to_numpy().tolist()  # Create list of list using rows
    tree["height"] = 10  # Number of rows to display, default is 10
    tree["show"] = "headings"
    # column identifiers
    tree["columns"] = l1
    for i in l1:
        tree.column(i, width=95, anchor="c")
        # Headings of respective columns
        tree.heading(i, text=i)
    for dt in r_set:
        row = [r for r in dt]  # creating a list from each row
        tree.insert("", "end", values=row)  # adding row
    style = ttk.Style()
    style.theme_use('clam')
    tree.place(bordermode=tk.INSIDE, relx=0.0, rely=0.71, anchor=tk.W,width=640,height=310)
    tree_scroll.place(bordermode=tk.INSIDE, relx=0.6, rely=0.71, anchor=tk.W,width=15,height=310,)

def clear_treeview():
    tree.delete(*tree.get_children())

def clicker(e):
    select_doc()

def select_doc():
    global jenvar, nomvar, pervar, tree, datjen, dattang,datper, dathal, datnom, datnam, datket
    
    row_id = tree.selection()
    select = tree.set(row_id)
    print(select)
    
    jenvar.set("Jenis : " + select['Jenis'])
    nomvar.set("Nomor : " + select['Nomor'])
    pervar.set("Perusahaan : " + select['Perusahaan'])
    
    datjen= (select['Jenis'])
    datnom= (select["Nomor"])
    datper= (select['Perusahaan'])
    dattang= (select['Tanggal'])
    dathal= (select['Perihal'])
    datket= (select['Keterangan'])
    datnam= (select['Nama'])

def invent():
    global datjen, dattang, datper, dathal, datnom, datnam, datket

    df_inv = pd.read_excel(r"C:\Users\ASUS\Documents\lemari\lemari_project\Book2.xlsx")
    
    # Use boolean indexing to filter the DataFrame
    filter_condition = (
        (df_inv["Jenis"] == datjen) &
        (df_inv["Nomor"] == datnom) &
        (df_inv["Perusahaan"] == datper) &
        (df_inv["Tanggal"] == dattang) &
        (df_inv["Perihal"] == dathal) &
        (df_inv["Keterangan"] == datket) &
        (df_inv["Nama"] == datnam)
    )
    df_abc = df_inv[filter_condition]

    rodi = df_abc.index.tolist()
    if df_abc.empty:
        print('DataFrame is empty!')
        tk.messagebox.showwarning(title="nodata", message="Maaf Arsip Tidak Tersedia")
    else:
        path_inv = r"C:\Users\ASUS\Documents\lemari\lemari_project\Book2.xlsx"
        wb_inv = openpyxl.load_workbook(path_inv)
        sheet = wb_inv.active

        # Convert the list of indices to a list of indices to remove
        idx_to_remove = [idx + 2 for idx in rodi]  # Add 2 to each index

        print("Maximum rows before removing: ", sheet.max_row)
    
        # Delete rows in reverse order to avoid shifting indices
        for idx in reversed(idx_to_remove):
            sheet.delete_rows(idx)
    
        wb_inv.save(path_inv)
        print("Maximum rows after removing: ", sheet.max_row)
        take_doc()


def take_doc() :
    global dattang, datjen, datper,dathal, time, tk, datnom, datnam,messagebox, datket, conn
    tgl, bln, thn = dattang.split('-')
    bln2 = ('a'+bln)
    bln3 = ('b'+bln)
    waktu = datetime.datetime.now()
    tanggal = waktu.strftime("%d-%m-%Y")
    keterangan = ("keluar")
    
    if datjen == ("INVOICE") and datper ==("PT. Bintang Utara") and datket==("masuk"): 
      
      name = ask_nonempty_string("Name", "What is your name?")
      exfil= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx"
      workbook = openpyxl.load_workbook(exfil)
      sheet = workbook.active
      sheet.append([datjen,datnom,datper,tanggal,keterangan,name,dathal])
      workbook.save(exfil)
      
      mycursor= conn.cursor()
      sendbase = "UPDATE statusled SET Stat=%s WHERE ID=0"
      val=(bln,)
      mycursor.execute(sendbase,val)
      conn.commit()
      print("{} Data Berhasil Dimasukan".format(mycursor.rowcount))
      
      confirm_out()
      
      
    elif datjen == ("INVOICE") and datper !=("PT. Bintang Utara") and datket == ("masuk"):
        
        name = ask_nonempty_string("Name", "What is your name?")
        exfil= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx"
        workbook = openpyxl.load_workbook(exfil)
        sheet = workbook.active
        sheet.append([datjen,datnom,datper,tanggal,keterangan,name,dathal])
        workbook.save(exfil)
        
        mycursor= conn.cursor()
        sendbase = "UPDATE statusled SET Stat=%s WHERE ID=0"
        val=(bln2,)
        mycursor.execute(sendbase,val)
        conn.commit()
        print("{} Data Berhasil Dimasukan".format(mycursor.rowcount))
        
        confirm_out()
    
    elif datjen ==  ("PERIZINAN") and datket == ("masuk"):
        
        name = askstring('Name', 'What is your name?')
        exfil= r"C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx"
        workbook = openpyxl.load_workbook(exfil)
        sheet = workbook.active
        sheet.append([datjen,datnom,datper,tanggal,keterangan,name,dathal])
        workbook.save(exfil)
        
        mycursor= conn.cursor()
        sendbase = "UPDATE statusled SET Stat=%s WHERE ID=0"
        val=(bln3,)
        mycursor.execute(sendbase,val)
        conn.commit()
        print("{} Data Berhasil Dimasukan".format(mycursor.rowcount))
        
        confirm_out()
    
    if (datjen == ("PERIZINAN") and datket==("keluar")) or (datjen == ("INVOICE") and datket == ("keluar")):
        tk.messagebox.showwarning(title="Error", message="Document has been put out by : "+ datnam)
    

def confirm_out():
    global conn
    msg_box = tk.messagebox.askquestion('Drawer Check', 'Are you sure the document has been put out ?',
                                        icon='warning')
    if msg_box == 'yes':
        mycursor= conn.cursor()
        sendbase = "UPDATE statusled SET Stat='99' WHERE ID=0"
        mycursor.execute(sendbase)
        conn.commit()
        tk.messagebox.showwarning(title="Success", message="Arsip telah dicatat")
        open_file()


def exlog():
    global tree
    cal = Calendar(mainWindow, selectmode='day', date_pattern="dd-mm-y")
    cal.place(x=660,y=90)
    button_datepik.place_forget()
    def grad_date():
        global l1 
        ldate.config(text="Selected Date is: " + cal.get_date())
        ttgl=str(cal.get_date()).strip()
        cal.place_forget()
        btn_seldate.place_forget()
        
        clear_treeview()
        df = pd.read_excel(r'C:\Users\ASUS\Documents\lemari\lemari_project\Book1.xlsx')
        l1 = list(df)  # List of column names as headers
        query = ttgl # get user entered string
    
        str1 = df.Tanggal.str.contains(query, case= False) # name column value matching
            
        df2 = df[( str1 )]  # combine all conditions using | operator
        r_set = df2.to_numpy().tolist()  # Create list of list using rows
        tree["height"] = 20  # Number of rows to display, default is 10
        tree["show"] = "headings"
        # column identifiers
        tree["columns"] = l1
        for i in l1:
            tree.column(i, width=95, anchor="c")
        # Headings of respective columns
            tree.heading(i, text=i)
        for dt in r_set:
            row = [r for r in dt]  # creating a list from each row
            tree.insert("", "end", values=row)  # adding row
        style = ttk.Style()
        style.theme_use('clam')
        tree.place(bordermode=tk.INSIDE, relx=0.0, rely=0.71, anchor=tk.W,width=640,height=310)
        tree_scroll.place(bordermode=tk.INSIDE, relx=0.6, rely=0.71, anchor=tk.W,width=15,height=310,)
        btn_logexp = tk.Button(mainWindow, text="Save", command=export_to_excel)
        btn_logexp.place(bordermode=tk.INSIDE, relx=0.90, rely=0.40, anchor=tk.E,height=35)
        button_datepik.place(bordermode=tk.INSIDE, relx=0.78, rely=0.5, anchor=tk.E, width=120, height=35)
    
    def export_to_excel():
        global l1
        file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
        if file_path:
            df_filtered = pd.DataFrame([tree.item(item, "values") for item in tree.get_children()],
                                       columns= l1)
            df_filtered.to_excel(file_path, index=False)
        tk.messagebox.showwarning(title="Export Log", message="Success Save")
        open_file()
         
    btn_seldate = tk.Button(mainWindow, text="Get Date", command=grad_date)
    btn_seldate.place(bordermode=tk.INSIDE, relx=0.93, rely=0.5, anchor=tk.E, width=140, height=35)

    ldate = Label(mainWindow, text="", bg='#fcd700')
    ldate.place(x=660,y=250)


def resume(event = 0):
    global button1, button2, button, lmain, cancel, button3

    cancel = False

    button1.place_forget()
    button2.place_forget()
    button3.place_forget()

    mainWindow.bind('<Return>', prompt_ok)
    button.place(bordermode=tk.INSIDE, relx=0.15, rely=0.4, anchor=tk.E, width=150, height=40)
    lmain.after(10, show_frame)
    var.set("Getting Photo")

def changeCam(event=0, nextCam=-1):
    global camIndex, cap, fileName

    if nextCam == -1:
        camIndex += 1
    else:
        camIndex = nextCam
    del(cap)
    cap = cv2.VideoCapture(camIndex)

    #try to get a frame, if it returns nothing
    success, frame = cap.read()
    if not success:
        camIndex = 0
        del(cap)
        cap = cv2.VideoCapture(camIndex)

    f = open(fileName, 'w')
    f.write(str(camIndex))
    f.close()

try:
    f = open(fileName, 'r')
    camIndex = int(f.readline())
except:
    camIndex = 0

cap = cv2.VideoCapture(camIndex)
capWidth = cap.get(3)
capHeight = cap.get(4)

success, frame = cap.read()
if not success:
    if camIndex == 0:
        print("Error, No webcam found!")
        sys.exit(1)
    else:
        changeCam(nextCam=0)
        success, frame = cap.read()
        if not success:
            print("Error, No webcam found!")
            sys.exit(1)


mainWindow = tk.Tk(screenName="Camera Capture")
mainWindow.wait_visibility()
mainWindow.geometry("1000x600")
mainWindow.resizable(width=False, height=False)
mainWindow.bind('<Escape>', lambda e: mainWindow.quit())
lmain = tk.Label(mainWindow, compound=tk.CENTER, anchor=tk.CENTER, relief=tk.RAISED)
button = tk.Button(mainWindow, text="Capture", command=prompt_ok)
button_changeCam = tk.Button(mainWindow, text="Switch Camera", command=changeCam)
button_search= tk.Button(mainWindow, text="Search",command = search)
button_take=tk.Button (mainWindow, text= "Take Doc" ,command= invent)
button_rest= tk.Button (mainWindow, text= "Reset" ,command = open_file)
button_datepik=tk.Button (mainWindow, text= "Export Log" ,command = exlog)
se_entry=tk.Entry(mainWindow)
tulisan = tk.Label(text ="Result", fg='black',font="Arial 9").place(x=470,y=5)
lsearch = tk.Label(text ="Search Document", fg='black',font="Arial 11").place(x=660,y=340)
var = tk.StringVar()
Hasocr = tk.Label(mainWindow, textvariable = var,
    fg='black',font="Arial 12",bg='white').place(x=470,y=30)
var.set("Wellcome.., Put your archieve in camera box")

jenvar = tk.StringVar()
nomvar = tk.StringVar()
pervar = tk.StringVar()
jenlab = tk.Label(mainWindow, textvariable = jenvar,
    fg='black',font="Arial 10",bg='#fcd700').place(x=660,y=460)
jenvar.set("Jenis :                   ")
nomlab = tk.Label(mainWindow, textvariable = nomvar,
    fg='black',font="Arial 10",bg='#fcd700').place(x=660,y=485)
nomvar.set("Nomor :                ")
perlab = tk.Label(mainWindow, textvariable = pervar,
    fg='black',font="Arial 10",bg='#fcd700').place(x=660,y=510)
pervar.set("Perusahaan :        ")


tree_scroll = Scrollbar(mainWindow)
# Create Treeview
tree = ttk.Treeview(mainWindow, yscrollcommand=tree_scroll.set, selectmode="extended")
tree_scroll.config(command = tree.yview)
open_file()


lmain.grid(row=0, column=0)
button.place(bordermode=tk.INSIDE, relx=0.15, rely=0.4, anchor=tk.E, width=150, height=40)
button.focus()
button_changeCam.place(bordermode=tk.INSIDE, relx=0.3, rely=0.4, anchor=tk.E, width=140, height=40)
button_datepik.place(bordermode=tk.INSIDE, relx=0.78, rely=0.5, anchor=tk.E, width=120, height=35)
button_search.place(bordermode=tk.INSIDE, relx=0.78, rely=0.69, anchor=tk.E, width=120, height=35)
button_rest.place(bordermode=tk.INSIDE, relx=0.93, rely=0.69, anchor=tk.E, width=120, height=35)
button_take.place (bordermode=tk.INSIDE, relx=0.78, rely=0.94, anchor=tk.E, width=120, height=35)
se_entry.place(bordermode=tk.INSIDE, relx=0.66, rely=0.63, anchor=tk.W, width=300, height=22)
tree.bind("<ButtonRelease-1>", clicker)

def show_frame():
    global cancel, prevImg, button

    _, frame = cap.read()
    cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
    rezim=cv2.resize(cv2image,(300,200))
    prevImg = Image.fromarray(cv2image)
    prevImg2 = Image.fromarray(rezim)
    imgtk = ImageTk.PhotoImage(image=prevImg2)
    lmain.imgtk = imgtk
    lmain.configure(image=imgtk)
    if not cancel:
        lmain.after(10, show_frame)

show_frame()
mainWindow.mainloop()

